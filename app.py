import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import sqlite3
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter


class LigaTenisMesa:
    def __init__(self, root):
        self.root = root
        self.root.title("Liga de Tenis de Mesa - Round Robin")
        self.root.geometry("1200x800")

        # Conectar a la base de datos
        self.conn = sqlite3.connect('liga_tenis_mesa.db')
        self.c = self.conn.cursor()

        # Crear tablas si no existen
        self.crear_tablas()

        # Variables de control
        self.jugadores = []
        self.partidos_jornada_actual = []
        self.jornada_actual = 1
        self.cargar_jornada_actual()

        # Configurar la interfaz
        self.configurar_interfaz()

        # Cargar datos iniciales
        self.actualizar_lista_jugadores()
        self.actualizar_tabla_ranking()
        self.generar_partidos_jornada()

    def crear_tablas(self):
        # Tabla de jugadores
        self.c.execute('''CREATE TABLE IF NOT EXISTS jugadores (
                          id INTEGER PRIMARY KEY AUTOINCREMENT,
                          nombre TEXT UNIQUE,
                          elo INTEGER DEFAULT 1000,
                          partidos_jugados INTEGER DEFAULT 0,
                          partidos_ganados INTEGER DEFAULT 0,
                          partidos_perdidos INTEGER DEFAULT 0,
                          puntos INTEGER DEFAULT 0
                          )''')

        # Tabla de partidos
        self.c.execute('''CREATE TABLE IF NOT EXISTS partidos (
                          id INTEGER PRIMARY KEY AUTOINCREMENT,
                          jornada INTEGER,
                          jugador1_id INTEGER,
                          jugador2_id INTEGER,
                          ganador_id INTEGER,
                          fecha TEXT,
                          FOREIGN KEY(jugador1_id) REFERENCES jugadores(id),
                          FOREIGN KEY(jugador2_id) REFERENCES jugadores(id),
                          FOREIGN KEY(ganador_id) REFERENCES jugadores(id)
                          )''')

        # Tabla de jornadas
        self.c.execute('''CREATE TABLE IF NOT EXISTS jornadas (
                          numero INTEGER PRIMARY KEY,
                          completada INTEGER DEFAULT 0
                          )''')

        self.conn.commit()

    def configurar_interfaz(self):
        # Frame principal
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Sección de jugadores
        jugadores_frame = ttk.LabelFrame(main_frame, text="Jugadores", padding="10")
        jugadores_frame.grid(row=0, column=0, padx=5, pady=5, sticky="nsew")

        self.lista_jugadores = ttk.Treeview(jugadores_frame, columns=('nombre', 'elo'), show='headings')
        self.lista_jugadores.heading('nombre', text='Nombre')
        self.lista_jugadores.heading('elo', text='ELO')
        self.lista_jugadores.column('nombre', width=150)
        self.lista_jugadores.column('elo', width=50)
        self.lista_jugadores.pack(fill=tk.BOTH, expand=True)

        # Controles de jugadores
        jugadores_controls = ttk.Frame(jugadores_frame)
        jugadores_controls.pack(fill=tk.X, pady=5)

        self.nuevo_jugador_entry = ttk.Entry(jugadores_controls)
        self.nuevo_jugador_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)

        ttk.Button(jugadores_controls, text="Añadir Jugador",
                   command=self.agregar_jugador).pack(side=tk.LEFT, padx=5)

        # Sección de ranking
        ranking_frame = ttk.LabelFrame(main_frame, text="Ranking", padding="10")
        ranking_frame.grid(row=1, column=0, padx=5, pady=5, sticky="nsew")

        self.tabla_ranking = ttk.Treeview(ranking_frame,
                                          columns=('posicion', 'nombre', 'elo', 'pj', 'pg', 'pp', 'puntos'),
                                          show='headings')
        self.tabla_ranking.heading('posicion', text='Pos')
        self.tabla_ranking.heading('nombre', text='Nombre')
        self.tabla_ranking.heading('elo', text='ELO')
        self.tabla_ranking.heading('pj', text='PJ')
        self.tabla_ranking.heading('pg', text='PG')
        self.tabla_ranking.heading('pp', text='PP')
        self.tabla_ranking.heading('puntos', text='Puntos')

        for col in ('posicion', 'elo', 'pj', 'pg', 'pp', 'puntos'):
            self.tabla_ranking.column(col, width=50, anchor='center')
        self.tabla_ranking.column('nombre', width=150)

        self.tabla_ranking.pack(fill=tk.BOTH, expand=True)

        # Sección de partidos
        partidos_frame = ttk.LabelFrame(main_frame, text=f"Jornada {self.jornada_actual}", padding="10")
        partidos_frame.grid(row=0, column=1, rowspan=2, padx=5, pady=5, sticky="nsew")

        self.tabla_partidos = ttk.Treeview(partidos_frame, columns=('jugador1', 'jugador2', 'ganador'), show='headings')
        self.tabla_partidos.heading('jugador1', text='Jugador 1')
        self.tabla_partidos.heading('jugador2', text='Jugador 2')
        self.tabla_partidos.heading('ganador', text='Ganador')

        for col in ('jugador1', 'jugador2', 'ganador'):
            self.tabla_partidos.column(col, width=150)

        self.tabla_partidos.pack(fill=tk.BOTH, expand=True)

        # Controles de partidos
        partidos_controls = ttk.Frame(partidos_frame)
        partidos_controls.pack(fill=tk.X, pady=5)

        ttk.Button(partidos_controls, text="Registrar Resultado",
                   command=self.registrar_resultado).pack(side=tk.LEFT, padx=5)

        ttk.Button(partidos_controls, text="Importar desde Excel",
                   command=self.importar_desde_excel).pack(side=tk.LEFT, padx=5)

        ttk.Button(partidos_controls, text="Finalizar Jornada",
                   command=self.finalizar_jornada).pack(side=tk.LEFT, padx=5)

        # Configurar pesos de filas y columnas
        main_frame.columnconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(0, weight=1)
        main_frame.rowconfigure(1, weight=1)

    def cargar_jornada_actual(self):
        self.c.execute("SELECT MAX(numero) FROM jornadas")
        result = self.c.fetchone()
        if result and result[0]:
            self.jornada_actual = result[0]
            self.c.execute("SELECT completada FROM jornadas WHERE numero=?", (self.jornada_actual,))
            completada = self.c.fetchone()[0]
            if completada:
                self.jornada_actual += 1
        else:
            # Primera ejecución, insertar jornada inicial
            self.jornada_actual = 1
            self.c.execute("INSERT INTO jornadas (numero, completada) VALUES (?, ?)", (self.jornada_actual, 0))
            self.conn.commit()

    def actualizar_lista_jugadores(self):
        # Limpiar lista
        for item in self.lista_jugadores.get_children():
            self.lista_jugadores.delete(item)

        # Obtener jugadores de la base de datos
        self.c.execute("SELECT nombre, elo FROM jugadores ORDER BY nombre")
        self.jugadores = [{'nombre': row[0], 'elo': row[1]} for row in self.c.fetchall()]

        # Agregar a la lista
        for jugador in self.jugadores:
            self.lista_jugadores.insert('', 'end', values=(jugador['nombre'], jugador['elo']))

    def actualizar_tabla_ranking(self):
        # Limpiar tabla
        for item in self.tabla_ranking.get_children():
            self.tabla_ranking.delete(item)

        # Obtener ranking de la base de datos
        self.c.execute('''SELECT nombre, elo, partidos_jugados, partidos_ganados, 
                         partidos_perdidos, puntos FROM jugadores 
                         ORDER BY puntos DESC, elo DESC, partidos_ganados DESC''')
        ranking = self.c.fetchall()

        # Agregar a la tabla
        for i, row in enumerate(ranking, 1):
            self.tabla_ranking.insert('', 'end', values=(i, row[0], row[1], row[2], row[3], row[4], row[5]))

    def generar_partidos_jornada(self):
        # Limpiar tabla de partidos
        for item in self.tabla_partidos.get_children():
            self.tabla_partidos.delete(item)

        # Verificar si ya hay partidos generados para esta jornada
        self.c.execute('''SELECT j1.nombre, j2.nombre, jg.nombre 
                         FROM partidos p
                         JOIN jugadores j1 ON p.jugador1_id = j1.id
                         JOIN jugadores j2 ON p.jugador2_id = j2.id
                         LEFT JOIN jugadores jg ON p.ganador_id = jg.id
                         WHERE p.jornada = ?''', (self.jornada_actual,))
        partidos_existentes = self.c.fetchall()

        if partidos_existentes:
            # Mostrar partidos existentes
            self.partidos_jornada_actual = []
            for partido in partidos_existentes:
                self.tabla_partidos.insert('', 'end', values=(partido[0], partido[1], partido[2] if partido[2] else ""))
                self.partidos_jornada_actual.append({
                    'jugador1': partido[0],
                    'jugador2': partido[1],
                    'ganador': partido[2] if partido[2] else None
                })
        else:
            # Generar nuevos partidos para la jornada
            self.partidos_jornada_actual = self.generar_round_robin()

            # Insertar partidos en la base de datos
            for partido in self.partidos_jornada_actual:
                # Obtener IDs de los jugadores
                self.c.execute("SELECT id FROM jugadores WHERE nombre=?", (partido['jugador1'],))
                jugador1_id = self.c.fetchone()[0]
                self.c.execute("SELECT id FROM jugadores WHERE nombre=?", (partido['jugador2'],))
                jugador2_id = self.c.fetchone()[0]

                self.c.execute('''INSERT INTO partidos 
                                (jornada, jugador1_id, jugador2_id, fecha)
                                VALUES (?, ?, ?, ?)''',
                               (self.jornada_actual, jugador1_id, jugador2_id, datetime.now().strftime("%Y-%m-%d")))
                self.conn.commit()

                # Mostrar en la tabla
                self.tabla_partidos.insert('', 'end', values=(partido['jugador1'], partido['jugador2'], ""))

    def generar_round_robin(self):
        if len(self.jugadores) < 2:
            return []

        # Algoritmo Round Robin para generar los partidos de la jornada
        jugadores = [j['nombre'] for j in self.jugadores]
        n = len(jugadores)
        partidos = []

        # Para cada jornada, cada jugador juega 3 partidos (contra 3 jugadores diferentes)
        # En total son 11 jornadas, cada jugador juega contra cada otro jugador exactamente 3 veces

        # Primero, verificar qué partidos ya se han jugado en jornadas anteriores
        self.c.execute('''SELECT j1.nombre, j2.nombre, COUNT(*) as veces_jugado
                         FROM partidos p
                         JOIN jugadores j1 ON p.jugador1_id = j1.id
                         JOIN jugadores j2 ON p.jugador2_id = j2.id
                         GROUP BY j1.nombre, j2.nombre''')
        partidos_anteriores = self.c.fetchall()

        # Crear un diccionario de enfrentamientos
        enfrentamientos = {}
        for j1 in jugadores:
            enfrentamientos[j1] = {}
            for j2 in jugadores:
                if j1 != j2:
                    enfrentamientos[j1][j2] = 0

        for partido in partidos_anteriores:
            j1, j2, count = partido
            enfrentamientos[j1][j2] = count
            enfrentamientos[j2][j1] = count

        # Seleccionar los partidos para esta jornada
        partidos_jornada = []
        jugadores_por_jugar = jugadores.copy()

        while len(partidos_jornada) < (n * 3) / 2 and len(jugadores_por_jugar) > 1:
            # Encontrar el jugador que ha jugado menos partidos en esta jornada
            jugador_actual = min(jugadores_por_jugar,
                                 key=lambda x: sum(1 for p in partidos_jornada if x in [p['jugador1'], p['jugador2']]))

            # Encontrar oponente con quien ha jugado menos veces y que no esté ya en 3 partidos esta jornada
            posibles_oponentes = [
                j for j in jugadores
                if j != jugador_actual and
                   j in jugadores_por_jugar and
                   sum(1 for p in partidos_jornada if j in [p['jugador1'], p['jugador2']]) < 3 and
                   (j, jugador_actual) not in [(p['jugador1'], p['jugador2']) for p in partidos_jornada] and
                   (jugador_actual, j) not in [(p['jugador1'], p['jugador2']) for p in partidos_jornada]
            ]

            if not posibles_oponentes:
                # No hay oponentes válidos, pasar al siguiente jugador
                jugadores_por_jugar.remove(jugador_actual)
                continue

            # Seleccionar el oponente con el que ha jugado menos veces
            oponente = min(posibles_oponentes, key=lambda x: enfrentamientos[jugador_actual][x])

            # Crear el partido
            partido = {
                'jugador1': jugador_actual,
                'jugador2': oponente,
                'ganador': None
            }
            partidos_jornada.append(partido)

            # Actualizar contadores de enfrentamientos
            enfrentamientos[jugador_actual][oponente] += 1
            enfrentamientos[oponente][jugador_actual] += 1

            # Verificar si algún jugador ya tiene 3 partidos en esta jornada
            for jugador in [jugador_actual, oponente]:
                if sum(1 for p in partidos_jornada if jugador in [p['jugador1'], p['jugador2']]) >= 3:
                    if jugador in jugadores_por_jugar:
                        jugadores_por_jugar.remove(jugador)

        return partidos_jornada

    def agregar_jugador(self):
        nombre = self.nuevo_jugador_entry.get().strip()
        if not nombre:
            messagebox.showerror("Error", "Por favor ingrese un nombre para el jugador")
            return

        try:
            self.c.execute("INSERT INTO jugadores (nombre) VALUES (?)", (nombre,))
            self.conn.commit()
            self.nuevo_jugador_entry.delete(0, tk.END)
            self.actualizar_lista_jugadores()
            self.actualizar_tabla_ranking()
            # Regenerar partidos si es necesario
            if self.jornada_actual == 1 and not self.partidos_jornada_actual:
                self.generar_partidos_jornada()
        except sqlite3.IntegrityError:
            messagebox.showerror("Error", f"El jugador '{nombre}' ya existe en la liga")

    def registrar_resultado(self):
        seleccion = self.tabla_partidos.selection()
        if not seleccion:
            messagebox.showerror("Error", "Por favor seleccione un partido para registrar el resultado")
            return

        item = seleccion[0]
        partido_idx = self.tabla_partidos.index(item)
        partido = self.partidos_jornada_actual[partido_idx]

        if partido['ganador'] is not None:
            messagebox.showerror("Error", "Este partido ya tiene un resultado registrado")
            return

        # Ventana de diálogo para seleccionar ganador
        dialog = tk.Toplevel(self.root)
        dialog.title("Registrar Resultado")
        dialog.geometry("300x150")

        ttk.Label(dialog, text=f"{partido['jugador1']} vs {partido['jugador2']}").pack(pady=10)

        ganador_var = tk.StringVar()
        ttk.Radiobutton(dialog, text=partido['jugador1'], variable=ganador_var, value=partido['jugador1']).pack()
        ttk.Radiobutton(dialog, text=partido['jugador2'], variable=ganador_var, value=partido['jugador2']).pack()

        def guardar_resultado():
            ganador = ganador_var.get()
            if not ganador:
                messagebox.showerror("Error", "Por favor seleccione un ganador")
                return

            self.registrar_resultado_partido(partido_idx, ganador)
            dialog.destroy()

        ttk.Button(dialog, text="Guardar", command=guardar_resultado).pack(pady=10)

    def registrar_resultado_partido(self, partido_idx, ganador):
        partido = self.partidos_jornada_actual[partido_idx]

        # Actualizar en memoria
        partido['ganador'] = ganador
        self.partidos_jornada_actual[partido_idx] = partido

        # Actualizar en la base de datos
        self.c.execute('''UPDATE partidos SET ganador_id = 
                        (SELECT id FROM jugadores WHERE nombre = ?)
                        WHERE jornada = ? AND 
                        jugador1_id = (SELECT id FROM jugadores WHERE nombre = ?) AND
                        jugador2_id = (SELECT id FROM jugadores WHERE nombre = ?)''',
                       (ganador, self.jornada_actual, partido['jugador1'], partido['jugador2']))

        # Actualizar estadísticas de los jugadores
        perdedor = partido['jugador2'] if ganador == partido['jugador1'] else partido['jugador1']

        # Actualizar ganador
        self.c.execute('''UPDATE jugadores SET 
                        partidos_jugados = partidos_jugados + 1,
                        partidos_ganados = partidos_ganados + 1,
                        puntos = puntos + 3,
                        elo = elo + 20
                        WHERE nombre = ?''', (ganador,))

        # Actualizar perdedor
        self.c.execute('''UPDATE jugadores SET 
                        partidos_jugados = partidos_jugados + 1,
                        partidos_perdidos = partidos_perdidos + 1,
                        elo = elo - 10
                        WHERE nombre = ?''', (perdedor,))

        self.conn.commit()

        # Actualizar visualización
        item = self.tabla_partidos.get_children()[partido_idx]
        self.tabla_partidos.item(item, values=(partido['jugador1'], partido['jugador2'], ganador))
        self.actualizar_tabla_ranking()

    def importar_desde_excel(self):
        # Pedir al usuario que seleccione el archivo Excel
        filepath = filedialog.askopenfilename(
            title="Seleccionar archivo Excel",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )

        if not filepath:
            return  # Usuario canceló

        try:
            # Leer el archivo Excel
            workbook = openpyxl.load_workbook(filepath)
            sheet = workbook.active

            # Buscar los partidos en el Excel
            partidos_encontrados = []

            for row in sheet.iter_rows(values_only=True):
                if len(row) >= 3 and row[0] and row[1]:  # Jugador1 y Jugador2 existen
                    jugador1 = str(row[0]).strip()
                    jugador2 = str(row[1]).strip()
                    ganador = str(row[2]).strip() if len(row) > 2 and row[2] else None

                    if jugador1 and jugador2:
                        partidos_encontrados.append({
                            'jugador1': jugador1,
                            'jugador2': jugador2,
                            'ganador': ganador
                        })

            if not partidos_encontrados:
                messagebox.showerror("Error", "No se encontraron partidos en el archivo Excel")
                return

            # Verificar que todos los jugadores existen en la base de datos
            jugadores_existentes = {j['nombre'] for j in self.jugadores}
            jugadores_en_excel = set()

            for partido in partidos_encontrados:
                jugadores_en_excel.add(partido['jugador1'])
                jugadores_en_excel.add(partido['jugador2'])
                if partido['ganador']:
                    jugadores_en_excel.add(partido['ganador'])

            jugadores_faltantes = jugadores_en_excel - jugadores_existentes
            if jugadores_faltantes:
                messagebox.showerror("Error",
                                     f"Los siguientes jugadores no existen en la liga:\n{', '.join(jugadores_faltantes)}")
                return

            # Procesar los partidos encontrados
            partidos_importados = 0

            for partido_excel in partidos_encontrados:
                # Buscar el partido correspondiente en la jornada actual
                for i, partido in enumerate(self.partidos_jornada_actual):
                    if ((partido['jugador1'] == partido_excel['jugador1'] and
                         partido['jugador2'] == partido_excel['jugador2']) or
                            (partido['jugador1'] == partido_excel['jugador2'] and
                             partido['jugador2'] == partido_excel['jugador1'])):

                        if partido['ganador'] is not None:
                            continue  # Ya tiene resultado

                        if partido_excel['ganador']:
                            # Registrar el resultado
                            self.registrar_resultado_partido(i, partido_excel['ganador'])
                            partidos_importados += 1
                        break

            messagebox.showinfo("Importación completada",
                                f"Se importaron resultados para {partidos_importados} partidos")

        except Exception as e:
            messagebox.showerror("Error", f"No se pudo leer el archivo Excel:\n{str(e)}")

    def finalizar_jornada(self):
        # Verificar que todos los partidos tienen resultado
        for partido in self.partidos_jornada_actual:
            if partido['ganador'] is None:
                messagebox.showerror("Error", "Hay partidos sin resultado registrado")
                return

        # Marcar jornada como completada
        self.c.execute("UPDATE jornadas SET completada = 1 WHERE numero = ?", (self.jornada_actual,))

        # Insertar nueva jornada si no es la última
        if self.jornada_actual < 11:
            self.jornada_actual += 1
            self.c.execute("INSERT INTO jornadas (numero, completada) VALUES (?, ?)", (self.jornada_actual, 0))
            self.conn.commit()

            # Actualizar interfaz
            partidos_frame = self.tabla_partidos.master
            partidos_frame.config(text=f"Jornada {self.jornada_actual}")
            self.generar_partidos_jornada()

            messagebox.showinfo("Éxito",
                                f"Jornada {self.jornada_actual - 1} completada. Jornada {self.jornada_actual} generada.")
        else:
            messagebox.showinfo("Liga Completa", "¡Todas las jornadas han sido completadas!")

    def __del__(self):
        self.conn.close()


if __name__ == "__main__":
    root = tk.Tk()
    app = LigaTenisMesa(root)
    root.mainloop()