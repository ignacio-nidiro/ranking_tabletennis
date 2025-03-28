import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def leer_archivo_excel(archivo_entrada):
    """
    Lee un archivo Excel y extrae datos de diferentes columnas
    """
    try:
        # Cargar el libro de trabajo existente
        libro = openpyxl.load_workbook(archivo_entrada)
        hoja = libro['Jugadores']

        datos = []
        encabezados = []

        # Leer encabezados (primera fila)
        for col in range(1, hoja.max_column + 1):
            encabezados.append(hoja.cell(row=1, column=col).value)

        # Leer datos de las filas siguientes
        for fila in range(2, hoja.max_row + 1):
            fila_datos = {}
            for col in range(1, hoja.max_column + 1):
                nombre_columna = encabezados[col - 1]
                valor_celda = hoja.cell(row=fila, column=col).value
                fila_datos[nombre_columna] = valor_celda
            datos.append(fila_datos)

        print("Datos leídos correctamente:")
        for dato in datos:  # Mostrar solo los primeros 3 registros
            print(dato)

        return datos

    except Exception as e:
        print(f"Error al leer el archivo: {e}")
        return None


def crear_archivo_excel(datos, archivo_salida):
    """
    Crea un nuevo archivo Excel con formato específico
    """
    try:
        # Crear un nuevo libro de trabajo
        libro = openpyxl.load_workbook(archivo_salida)
        try:
            hoja = libro['Rol de Liga']
        except Exception:
            libro.create_sheet('Rol de Liga')
            hoja = libro['Rol de Liga']

        hoja.title = "Rol de Liga"

        # Definir estilos
        estilo_encabezado = Font(name='Arial', bold=True, size=12, color='FFFFFF')
        alineacion_centro = Alignment(horizontal='center', vertical='center')
        relleno_encabezado = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        borde_fino = Border(left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin'))

        # Escribir encabezados
        encabezados = list(datos[0].keys()) if datos else []
        print(encabezados)
        for y in range(1, 13):
            for x in range(1, 4):
                encabezados.append(f'{y} Match#{x}')
        for col, encabezado in enumerate(encabezados, start=1):
            celda = hoja.cell(row=1, column=col, value=encabezado)
            celda.font = estilo_encabezado
            celda.alignment = alineacion_centro
            celda.fill = relleno_encabezado
            celda.border = borde_fino

        # Escribir datos
        for fila, dato in enumerate(datos, start=2):
            for col, (key, value) in enumerate(dato.items(), start=1):
                celda = hoja.cell(row=fila, column=col, value=value)
                celda.border = borde_fino

                if isinstance(value, (int, float)):
                    celda.alignment = Alignment(horizontal='right')
                else:
                    celda.alignment = Alignment(horizontal='left')


        # Ajustar el ancho de las columnas al contenido y ponerles borde
        for col in range(1, len(encabezados) + 1):
            col_letra = get_column_letter(col)
            hoja.column_dimensions[col_letra].width = max(15, len(str(encabezados[col - 1])) * 1.2)
            for fila, _ in enumerate(datos, start=2):
                hoja.cell(fila, col).border = borde_fino

        # Pintar celdas de partidos propios
        for col, encabezado in enumerate(encabezados[2:], start=3):
            for fila, dato in enumerate(datos, start=1):
                if int(encabezado.split(' ')[0]) == dato['No. Jugador']:
                    hoja.cell(fila+1, col).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type="solid")

        # Guardar el archivo
        libro.save(archivo_salida)
        print(f"Archivo '{archivo_salida}' creado exitosamente con formato.")

    except Exception as e:
        print(f"Error al crear el archivo: {e}")
        raise e


# Ejemplo de uso
if __name__ == "__main__":
    # Nombre del archivo de entrada y salida
    archivo = 'TablaRoundRobin.xlsx'

    # 1. Leer el archivo Excel original
    datos = leer_archivo_excel(archivo)

    if datos:
        # 2. Procesar los datos (ejemplo: agregar una columna calculada)
        for dato in datos:
            if 'Precio' in dato and 'Cantidad' in dato:
                dato['Total'] = dato['Precio'] * dato['Cantidad']

        # 3. Crear nuevo archivo con formato
        crear_archivo_excel(datos, archivo)